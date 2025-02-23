import os
import requests
import re
import sqlite3
import openpyxl
from tkinter import filedialog
from openpyxl.worksheet.worksheet import Worksheet
from tkinter import messagebox
from openpyxl.utils.exceptions import InvalidFileException
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import UnexpectedAlertPresentException, NoAlertPresentException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import random



def buscador_de_frases(palavra, contador_offset):
    try:
        with sqlite3.connect(r'arquivos_extras\frases.db') as conexao:
            cursor = conexao.cursor()
            cursor.execute('''
            SELECT texto
            FROM frases_fts
            WHERE texto MATCH ?
            LIMIT 10
            OFFSET ?
            ''', (palavra, contador_offset))
            resultado_frases = cursor.fetchmany(10)

            frases_encontradas = []
            for frase in resultado_frases:
                frases_encontradas.append(frase[0])
            print(frases_encontradas)
            return frases_encontradas
    except sqlite3.Error as e:
        print(e)
    except Exception as error:
        print(error)

def iniciar_driver():
    chrome_options = Options()

    USER_AGENTS = ['Mozilla/5.0 (iPad; CPU OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5355d Safari/8536.25',
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                   "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
                   "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:115.0) Gecko/20100101 Firefox/115.0",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:114.0) Gecko/20100101 Firefox/114.0",
                   "Mozilla/5.0 (Macintosh; Intel Mac OS X 11_3) AppleWebKit/537.36 (KHTML, like Gecko) Version/14.1 Safari/537.36",
                   "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/537.36",
                   "Mozilla/5.0 (Linux; Android 12; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36",
                   "Mozilla/5.0 (Linux; Android 10; SM-A205U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/119.0.0.0 Safari/537.36",
                   ]

    user_agent = random.choice(USER_AGENTS)

    arguments = ['--lang=pt-BR', '--window-size=800,800', '--incognito', '--headless=new',
                 f'--user-agent={user_agent}', '--disable-blink-features=AutomationControlled',
                 '--disable-images', '--disable-webgl', '--disable-gpu', '--disable-dev-shm-usage',
                 '--no-sandbox', '--disable-extensions', '--disable-software-rasterizer']

    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1,
    })
    driver = webdriver.Chrome(options=chrome_options)

    return driver


def tradutor_de_palavras(palavras_a_traduzir: list):
    # conteudo_pagina = driver.page_source
    # Expressão regular para capturar o conteúdo dentro do <span class="display-term">
    # padrao_palavras = r'<span class="display-term">(.*?)</span>'
    # palavras = re.findall(padrao_palavras, conteudo_pagina)

    driver = iniciar_driver()

    palavras_traduzidas = {}
    for palavra in palavras_a_traduzir:
        while True:
            try:
                driver.get(f'https://context.reverso.net/traducao/ingles-portugues/{palavra}')

                try:
                    alert = driver.switch_to.alert
                    alert_text = alert.text
                    print(f"Alerta encontrado: {alert_text}")
                    alert.accept()  # Aceita o alerta
                    print('############  ACEITANDO O ALERTA  ############')
                except NoAlertPresentException:
                    pass  # Não há alerta presente

                driver.execute_script("window.alert = function() {};")  # Desabilita alertas
                driver.execute_script("window.confirm = function() { return true; };")  # Confirma automaticamente
                driver.execute_script("window.prompt = function() { return null; };")  # Ignora prompts

                # Espera até que os elementos estejam presentes no DOM
                wait = WebDriverWait(driver, 3)  # Timeout de 10 segundos
                elementos = wait.until(
                    EC.presence_of_all_elements_located((By.XPATH, '//div[@id="translations-content"]//span[@class="display-term"]')))
                driver.execute_script("window.stop();")

                palavras = []
                for indice, elemento in enumerate(elementos):
                    palavras.append(elemento.text)
                    if indice == 3:
                        break

                palavras_traduzidas[palavra] = palavras
                print(palavras_traduzidas)
                sleep(1)
                break
            except Exception as error:
                print(error)

    print(palavras_traduzidas)
    driver.quit()
    return palavras_traduzidas


class GerenciadorPlanilha:
    def __init__(self):
        self.planilha = openpyxl.Workbook()
        self.sheet: Worksheet = self.planilha['Sheet']
        self.sheet.append(['Frase', 'Palavra', 'Tradução'])

    def adicionar_dados(self, frase, palavra, significados):
        self.sheet.append([frase, palavra, significados])

    def salvar_planilha(self):
        caminho = filedialog.asksaveasfilename(
            title='Escolha o local para salvar',
            filetypes=[('Arquivo de planilha', '*.xlsx')],
            initialfile='Planilha Anki',
            defaultextension='.xlsx',
            confirmoverwrite=True
        )

        try:
            self.planilha.save(caminho)
            print(f"Planilha salva com sucesso em: {caminho}")
        except FileNotFoundError:
            messagebox.showwarning(title='Ops', message='Caminho inválido ou operação cancelada.')
        except Exception as e:
            print(f"Erro ao salvar a planilha: {e}")


def adicionar_cartao(baralho, frase, significado_palavra, palavra):
    palavra_escapada = re.escape(palavra)  # Protege caracteres especiais
    # Regex para permitir espaços, e destacar a expressão ou palavra, preservando a formatação
    frase_formatada = re.sub(rf'(?<!\w)({palavra_escapada})(?!\w)', r'<b>\1</b>', frase, flags=re.IGNORECASE)

    requisicao = {
        "action": "addNote",
        "version": 6,
        "params": {
            "note": {
                "deckName": f"{baralho}",
                "modelName": "Básico",
                "fields": {
                    "Frente": f"{frase_formatada}",
                    "Verso": f"{significado_palavra}"
                },
                "audio": [{
                    "url": f"https://translate.google.com/translate_tts?ie=UTF-8&client=tw-ob&tl=en&q={frase.replace(' ', '+')}",
                    "filename": f"{palavra}.mp3",
                    "fields": [
                        "Frente"
                    ]
                }],
                "options": {
                    "allowDuplicate": True
                }
                }
        }
    }
    try:
        resposta = requests.post(url='http://127.0.0.1:8765', json=requisicao)
        print(resposta.json())
    except requests.exceptions.ConnectionError:
        messagebox.showerror(title='Erro de conexão com o Anki',
                             message='Verifique se o Anki está aberto e verifique sua conexão com a internet.')
    else:
        return resposta.json()

def pegar_baralhos():
    requisicao = {
        "action": "deckNames",
        "version": 6
    }
    try:
        resposta = requests.get(url='http://127.0.0.1:8765', json=requisicao)
        print(resposta.json()['result'])
    except InvalidFileException:
        messagebox.showwarning(title='Atenção!',
                               message='Arquivo inválido ou vazio.')
    except requests.exceptions.ConnectionError:
        messagebox.showerror(title='Erro de conexão com o Anki',
                             message='Verifique se o Anki está aberto e verifique sua conexão com a internet.')
    else:
        return resposta.json()['result']


def escrever_log(campo_log, mensagem):
    campo_log.configure(state='normal')
    campo_log.insert('end', mensagem + os.linesep)
    campo_log.see('end')
    campo_log.configure(state='disabled')


def automatizar_anki(arquivo, baralho, campo_log):
    campo_log = campo_log
    qtd_frases = 0
    try:
        workbook = openpyxl.load_workbook(arquivo)
        # talvez possa usar workbook.active aqui
        sheet = workbook['Sheet']

        for linha in sheet.iter_rows(min_row=2, min_col=1):
            frase = linha[0].value
            palavra = linha[1].value
            traducao = linha[2].value
            traducao_palavra = f'{palavra} = {traducao}'

            resultado_requisicao = adicionar_cartao(baralho=baralho, frase=frase, significado_palavra=traducao_palavra, palavra=palavra)

            if resultado_requisicao['error'] is None:
                qtd_frases += 1
                escrever_log(campo_log=campo_log,
                             mensagem=f'✔ Cartão referente a "{palavra}" inserido com sucesso!')
    except Exception as e:
        escrever_log(campo_log=campo_log,
                     mensagem=f'Aconteceu algum erro na automação Anki. Informe o erro ao desenvolvedor: {e}')
    finally:
        escrever_log(campo_log=campo_log,
                     mensagem=f'A sua automação terminou. Foram adicionadas {qtd_frases} frases!')


if __name__ == '__main__':
    tradutor_de_palavras(['pie', 'piece', 'gold', 'silver'])
